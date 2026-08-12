# Builds the CareerPilot manual QA workbook.
# Every navigation path here is a real route taken from client/src/App.tsx, and every gate
# described is one that actually exists in the server code — a test plan that describes
# screens that do not exist is worse than no test plan.
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "CareerPilot-Manual-Test-Plan.xlsx")

INK   = "1F2937"
HEAD  = "1D4ED8"
SUBHD = "E8EEFB"
BAND  = "F7F9FC"
POS   = "EFFAF3"
NEG   = "FDF1F1"
COR   = "FEF8EC"
WHITE = "FFFFFF"

thin = Side(style="thin", color="D8DEE9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, row, ncols, fill=HEAD, color=WHITE, size=10):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color=color, size=size, name="Segoe UI")
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 30

def widths(ws, w):
    for i, x in enumerate(w, start=1):
        ws.column_dimensions[get_column_letter(i)].width = x

# ─────────────────────────────────────────────────────────────────────────────
# FEATURES — the "what / why / when / where" the QA team needs before testing
# ─────────────────────────────────────────────────────────────────────────────
FEATURES = [
    ("Student", "Join, OTP & Login",
     "Registration by mobile with a WhatsApp OTP, plus password login once set.",
     "It is the front door. Everything else in CareerPilot is gated on an identity, and a "
     "stranger has to become a known member without a salesperson touching them.",
     "First contact — from an ad, a college drive, or a shared passport card.",
     "/careerpilot/join · /careerpilot/login"),

    ("Student", "Career Readiness Assessment",
     "A free, deterministic (non-AI) assessment producing a career score, level and pathway.",
     "It is the hook and the qualifier. Free so anyone will start it; deterministic so it costs "
     "nothing to run at volume; scored so the paid roadmap feels earned rather than sold.",
     "Immediately after joining, before any payment is asked for.",
     "/careerpilot/assessment"),

    ("Student", "Membership & Payment",
     "₹1,599/year membership purchased through Razorpay, unlocking all paid entitlements.",
     "The only revenue event in the module. Everything upstream is designed to reach it and "
     "everything downstream is designed to justify it.",
     "When a student hits any paid feature, or from the result screen after the assessment.",
     "Any locked feature → Buy prompt → Razorpay checkout"),

    ("Student", "Mission Control & Streak",
     "A daily set of missions with a streak counter and a nightly close job.",
     "A yearly membership dies without a reason to open the app on a Tuesday. The streak is "
     "the retention mechanic that turns a purchase into a habit.",
     "Every day, as the member's landing surface.",
     "/careerpilot"),

    ("Student", "90-Day Roadmap",
     "A personalised plan derived from the assessment. 7 days free, the full 90 paid.",
     "Converts a score into a plan. The paywall sits exactly where the value becomes obvious "
     "— the student can see the shape of the plan before being asked to pay for the rest.",
     "Straight after the assessment result, and revisited weekly.",
     "/careerpilot/roadmap"),

    ("Student", "Practice — Coding, SQL, MCQ",
     "Runnable practice problems executed in a sandbox, with saved submissions.",
     "The daily habit that makes the membership feel used. Real execution, not multiple choice, "
     "because employers test code.",
     "Daily, usually alongside the day's mission.",
     "/careerpilot/practice · /careerpilot/practice/:id"),

    ("Student", "Resume Centre",
     "Build, import, AI-improve and AI-score a resume.",
     "The artefact a student actually leaves with. It is the most tangible proof that the "
     "membership produced something.",
     "Before applications, and after major milestones.",
     "/careerpilot/resume"),

    ("Student", "AI Mock Interview",
     "A spoken interview with an Indian-accented AI voice, short questions, stored feedback, "
     "optionally primed for a specific company.",
     "The headline feature and the main reason people pay. Interview nerves are the real "
     "blocker for most candidates, and rehearsal is the only cure.",
     "Once a student has a roadmap and some practice behind them; before real interviews.",
     "/careerpilot/interview"),

    ("Student", "Prepare Interviews — Company Hub",
     "Per-company pages with six tabs: Overview, Interview Pattern, Salary, Mock Test, "
     "Mock Interview and Questions. A company is only visible once it clears the readiness bar.",
     "Students do not prepare 'for interviews', they prepare for TCS. Company-specific "
     "preparation is what they actually search for.",
     "In the weeks before a campus drive or an application.",
     "/careerpilot/companies · /careerpilot/companies/:slug"),

    ("Student", "Company Mock Test",
     "A timed MCQ paper modelled on a company's real pattern, with a server-side clock.",
     "A rehearsal under the real clock. Knowing the questions is not the same as finishing "
     "in time, which is what actually eliminates candidates.",
     "After reading a company's pattern, before the real aptitude round.",
     "/careerpilot/companies/:slug → Mock Test → /careerpilot/mock-test/:id"),

    ("Student", "Tech News",
     "A short daily feed of technology and hiring news, drafted by AI from a pasted URL.",
     "A cheap daily reason to return, and it makes the product feel alive between missions.",
     "Daily, in passing.",
     "/careerpilot/news"),

    ("Student", "Coins, XP & Leaderboard",
     "XP for status, coins as currency, plus a global leaderboard and achievements.",
     "Motivation. XP marks progress that can never be taken away; coins are meant to be spent "
     "on rewards. NOTE FOR QA: the spending half is not built yet — coins can only be earned.",
     "Continuously, as a by-product of every other activity.",
     "/careerpilot/coins · /careerpilot/leaderboard · /careerpilot/achievements"),

    ("Student", "Public Passport Card",
     "A public, shareable, no-login profile card at a slug URL.",
     "Free distribution. A student shares their card and their friends discover the product "
     "without a rupee of ad spend.",
     "After a milestone worth showing off.",
     "/careerpilot/card/:slug"),

    ("Admin", "Plan & Entitlements Config",
     "Which features are free and which are paid, plus the membership price and duration.",
     "Pricing must be changeable without a deploy. Entitlements are stored as data so a "
     "commercial decision never waits on engineering.",
     "Whenever pricing or packaging changes.",
     "/admin/passport/config"),

    ("Admin", "Assessment Builder & Paper Design",
     "Questions, stage segments, scoring categories and the printed paper layout.",
     "The assessment is the funnel. If it cannot be edited by the team that owns conversion, "
     "the funnel cannot be improved.",
     "Whenever the assessment is tuned.",
     "/admin/passport/assessment · /admin/careerpilot/paper-design"),

    ("Admin", "Pathways & Missions",
     "The career tracks members are sorted into, and the daily mission content.",
     "Daily content has to be authored by a human on a human's schedule, not shipped in a release.",
     "Weekly content planning.",
     "/admin/passport/pathways · /admin/passport/missions"),

    ("Admin", "Members Admin",
     "Create, edit, activate and deactivate members; inspect their answers and interviews.",
     "Support and sales both need one person's entire history on one screen.",
     "On every support call.",
     "/admin/passport/students"),

    ("Admin", "Company Roster & Content",
     "Bulk-create companies, AI-draft their profiles, edit interview patterns, import or "
     "generate questions, and verify eligibility. Readiness is scored live.",
     "The content engine behind Prepare Interviews, and the deepest admin surface in the module. "
     "A company goes live automatically when it clears the bar — there is no publish button.",
     "Continuously, as the company library grows.",
     "/admin/passport/companies"),

    ("Admin", "Tech News Admin",
     "Paste a URL, have AI draft the item, edit and publish.",
     "Publishing has to take a minute or it will not happen daily.",
     "Daily.",
     "/admin/passport/news"),

    ("Admin", "Coins Admin",
     "Earning rules, caps, the coin config and the full ledger.",
     "Every coin number is admin-changeable by design, so the economy can be tuned against "
     "real behaviour rather than guesses.",
     "Whenever the economy is rebalanced.",
     "/admin/passport/coins"),

    ("Admin", "Staging Board",
     "Moves a member between career stages.",
     "Stage drives which missions and content a member sees; it must be correctable without "
     "touching the database.",
     "When a member's situation changes.",
     "/admin/careerpilot/staging"),
]

# ─────────────────────────────────────────────────────────────────────────────
# TEST CASES  (feature, type, priority, title, precondition, navigation, steps,
#              test data, expected result)
# ─────────────────────────────────────────────────────────────────────────────
T = []
def add(feature, ttype, prio, title, pre, nav, steps, data, exp):
    T.append([feature, ttype, prio, title, pre, nav, steps, data, exp])

F = "Join, OTP & Login"
add(F,"Positive","P0","Register a new member with a valid mobile","Mobile number never used on this tenant",
    "Open /careerpilot/join",
    "1. Open /careerpilot/join\n2. Enter full name, mobile and email\n3. Tap Send OTP\n4. Read the OTP from WhatsApp\n5. Enter it and tap Verify",
    "Name: Test Student\nMobile: 9876500001\nEmail: test1@example.com",
    "OTP arrives on WhatsApp within 60s. After verifying, the member lands on the assessment start screen and a member record exists in Admin > Members.")
add(F,"Positive","P1","Log in with OTP on a returning member","Member already registered",
    "Open /careerpilot/login",
    "1. Open /careerpilot/login\n2. Enter the registered mobile\n3. Tap Send OTP\n4. Enter the OTP",
    "Mobile: 9876500001",
    "Member is logged in and lands on /careerpilot with their existing progress intact — not a fresh assessment.")
add(F,"Positive","P1","Set a password and then log in with it","Member logged in via OTP, no password set",
    "/careerpilot → Profile → Set password",
    "1. Set a password of 8+ characters\n2. Log out\n3. Open /careerpilot/login\n4. Switch to password login\n5. Enter mobile and password",
    "Password: Test@2026",
    "Password is accepted, login succeeds without an OTP, and the session behaves identically to an OTP login.")
add(F,"Positive","P2","Resend OTP after the timer expires","OTP requested, countdown finished",
    "/careerpilot/join → OTP screen",
    "1. Request an OTP\n2. Wait for the resend timer to reach zero\n3. Tap Resend",
    "-",
    "A new OTP is sent and accepted. The previous OTP is rejected.")
add(F,"Negative","P0","Reject an invalid mobile number","-","/careerpilot/join",
    "1. Enter a 9-digit number\n2. Tap Send OTP",
    "Mobile: 987650000",
    "A clear inline message states the number must be a valid 10-digit Indian mobile. No OTP is sent and no member record is created.")
add(F,"Negative","P0","Reject a wrong OTP","OTP requested","/careerpilot/join → OTP screen",
    "1. Enter six digits that are not the OTP\n2. Tap Verify",
    "OTP: 000000",
    "Verification fails with a readable message. The member is NOT logged in and no session token is issued.")
add(F,"Negative","P1","Reject an expired OTP","OTP requested","/careerpilot/join → OTP screen",
    "1. Request an OTP\n2. Wait past the expiry window\n3. Enter the original OTP",
    "-",
    "Verification is refused as expired and the member is invited to request a new one.")
add(F,"Negative","P1","Reject a password under the minimum length","Member logged in","/careerpilot → Profile",
    "1. Enter a 5-character password\n2. Save",
    "Password: Ab1@x",
    "Save is refused with a message naming the 8-character minimum. No password is stored.")
add(F,"Corner","P1","Mobile entered with +91, spaces or dashes","-","/careerpilot/join",
    "1. Enter the same number in four formats\n2. Send OTP each time",
    "+91 98765 00001\n+919876500001\n098765-00001\n9876500001",
    "All four normalise to the SAME member. No duplicate accounts are created. (This was a real defect — three different normalisers existed.)")
add(F,"Corner","P2","Same mobile registers twice","Member already exists","/careerpilot/join",
    "1. Register with an existing mobile\n2. Complete the OTP",
    "Mobile: 9876500001",
    "The existing member is logged in rather than a duplicate being created. Progress is preserved.")
add(F,"Corner","P2","OTP requested many times in quick succession","-","/careerpilot/join",
    "1. Tap Send OTP repeatedly, as fast as the UI allows",
    "-",
    "Requests are throttled. The button disables or a cooldown message appears. WhatsApp is not spammed and cost is bounded.")

F = "Career Readiness Assessment"
add(F,"Positive","P0","Complete the assessment end to end","Newly registered member","/careerpilot/assessment",
    "1. Start the assessment\n2. Answer every question across all stage segments\n3. Submit",
    "Answer honestly as a final-year student",
    "A career score, level and pathway are shown. The result is reproducible — the same answers always give the same score, because scoring is deterministic.")
add(F,"Positive","P1","Resume a part-finished assessment","Assessment started, some answers given","/careerpilot/assessment",
    "1. Answer the first few questions\n2. Close the browser\n3. Log back in and reopen the assessment",
    "-",
    "Previously given answers are still selected and the member resumes where they stopped.")
add(F,"Positive","P1","View the result and the free roadmap preview","Assessment submitted","/careerpilot/assessment → Result",
    "1. Read the result screen\n2. Follow the roadmap link",
    "-",
    "Score, level, pathway and a 7-day preview are visible. The full 90-day plan is clearly marked as paid.")
add(F,"Negative","P0","Submit with unanswered questions","Assessment in progress","/careerpilot/assessment",
    "1. Leave several questions blank\n2. Tap Submit",
    "-",
    "Submission is blocked, or the blanks are clearly listed. The member is never scored on questions they never saw.")
add(F,"Negative","P1","Submit the assessment twice","Assessment already submitted","/careerpilot/assessment",
    "1. Reopen the assessment\n2. Attempt to submit again",
    "-",
    "A second submission does not overwrite the first silently. Either the result is shown, or a retake is offered explicitly.")
add(F,"Corner","P1","A category with no questions asked","Assessment configured so one scoring category has no active questions",
    "/careerpilot/assessment",
    "1. Complete the assessment\n2. Inspect the category breakdown on the result",
    "-",
    "The unasked category is ABSENT from the breakdown — not shown as zero. Scoring a member on something never asked is a defect that was previously fixed.")
add(F,"Corner","P2","Admin edits questions mid-attempt","Member has an assessment open","/admin/passport/assessment",
    "1. As admin, edit or remove a question\n2. As the member, submit the open attempt",
    "-",
    "The submission does not crash. The member is scored against what they were actually asked.")

F = "Membership & Payment"
add(F,"Positive","P0","Buy a membership end to end","Free member, price set to ₹1,599",
    "Any locked feature → Buy",
    "1. Tap a locked feature\n2. Tap Buy\n3. Complete Razorpay with a real card for ₹1\n4. Return to the app",
    "Use Razorpay test mode, or ₹1 live then refund",
    "Payment succeeds, entitlements unlock immediately, expiry is set one year out, and the payment appears in the admin record. NOTE: this path has never completed in production — test it first.")
add(F,"Positive","P0","The price shown is ₹1,599","Admin config priceInr = 1599","Any locked feature",
    "1. Open three different locked features\n2. Note the price on each prompt",
    "-",
    "Every prompt and the Razorpay order all read ₹1,599. KNOWN DEFECT: production currently holds 499 — this test should FAIL until config is corrected.")
add(F,"Positive","P1","Paid member sees every paid feature unlocked","Membership active","/careerpilot",
    "1. Visit roadmap, practice, resume, interview, companies and news in turn",
    "-",
    "No lock prompt appears anywhere. All paid entitlements resolve as unlocked.")
add(F,"Negative","P0","Abandoned payment leaves the member free","Free member","Buy → Razorpay",
    "1. Start checkout\n2. Close the Razorpay window without paying\n3. Return to the app",
    "-",
    "Nothing unlocks. The member remains free and no membership record is created.")
add(F,"Negative","P0","A tampered payment response is rejected","Free member, browser dev tools open","Buy → Razorpay",
    "1. Intercept the verify call\n2. Alter the signature or payment id\n3. Let it through",
    "-",
    "Server-side signature verification fails and NOTHING unlocks. A forged response must never grant a membership.")
add(F,"Corner","P1","Membership expiring today","Member whose expiresAt is today","/careerpilot",
    "1. Open a paid feature before and after the expiry moment",
    "-",
    "Access is allowed up to expiry and refused after. The prompt to renew is clear rather than an error.")
add(F,"Corner","P1","Buying while already a member","Active member","Buy prompt",
    "1. Reach a Buy prompt by direct URL\n2. Attempt to pay again",
    "-",
    "Either the purchase is prevented, or the existing membership is extended rather than replaced. The member must not lose remaining days.")
add(F,"Corner","P2","Payment succeeds but the app is closed mid-flow","Free member","Buy → Razorpay",
    "1. Pay successfully\n2. Kill the browser before returning\n3. Log in again",
    "-",
    "The membership is active on next login. Money taken must never leave the member unentitled.")

F = "Mission Control & Streak"
add(F,"Positive","P0","Complete today's missions","Paid member","/careerpilot",
    "1. Open the dashboard\n2. Complete each mission listed for today",
    "-",
    "Each mission marks complete, progress updates immediately, and coins/XP are credited per the admin rules.")
add(F,"Positive","P1","Streak increases across consecutive days","Member completed yesterday's missions","/careerpilot",
    "1. Complete today's missions\n2. Check the streak counter",
    "-",
    "The streak increments by exactly one. It does not reset.")
add(F,"Negative","P1","Free member cannot access daily missions","Free member","/careerpilot",
    "1. Open the dashboard",
    "-",
    "Missions are shown as locked with a clear upgrade prompt, not as an error or a blank panel.")
add(F,"Negative","P2","A mission cannot be completed twice for extra reward","Paid member","/careerpilot",
    "1. Complete a mission\n2. Replay the completion request",
    "-",
    "The second attempt credits nothing. The coin ledger's idempotency key prevents a double credit.")
add(F,"Corner","P1","A day is missed","Member with a streak who skips a day","/careerpilot",
    "1. Do not complete anything for a full day\n2. Return the following day",
    "-",
    "The streak resets according to the documented rule and the member is told why, rather than the number silently changing.")
add(F,"Corner","P2","Midnight crossing mid-session","Member active at 23:59 IST","/careerpilot",
    "1. Open missions at 23:58\n2. Complete one at 00:01",
    "-",
    "The completion is attributed to a single, unambiguous day. No mission is credited to both days or lost between them.")

F = "90-Day Roadmap"
add(F,"Positive","P0","Free member sees a 7-day preview","Assessment complete, free member","/careerpilot/roadmap",
    "1. Open the roadmap\n2. Scroll to the end of the visible plan",
    "-",
    "Exactly 7 days are readable, and the remainder is visibly present but locked with an upgrade prompt.")
add(F,"Positive","P0","Paid member sees the full 90 days","Membership active","/careerpilot/roadmap",
    "1. Open the roadmap\n2. Scroll to day 90",
    "-",
    "All 90 days are readable with no lock anywhere.")
add(F,"Negative","P1","Roadmap before the assessment","Member registered, assessment not taken","/careerpilot/roadmap",
    "1. Open the roadmap directly",
    "-",
    "The member is guided to take the assessment first. No empty or broken plan is rendered.")
add(F,"Corner","P2","Retaking the assessment changes the roadmap","Assessment already complete","/careerpilot/assessment",
    "1. Reset and retake with markedly different answers\n2. Reopen the roadmap",
    "-",
    "The roadmap reflects the new result. Stale content from the previous attempt does not survive.")

F = "Practice — Coding, SQL, MCQ"
add(F,"Positive","P0","Run correct code and pass","Paid member","/careerpilot/practice → open a problem",
    "1. Open a coding problem\n2. Write a correct solution\n3. Tap Run, then Submit",
    "A known-good solution in Python",
    "Output matches, all test cases pass, and the submission is saved and visible on return.")
add(F,"Positive","P1","Run Java successfully","Paid member","/careerpilot/practice/:id",
    "1. Choose Java\n2. Submit a correct solution",
    "Java Hello World plus the problem logic",
    "Compiles and runs within the timeout. Java is a heavy language and is throttled separately — it should still complete.")
add(F,"Positive","P2","Saved work survives a reload","Practice problem partly solved","/careerpilot/practice/:id",
    "1. Write half a solution\n2. Reload the page",
    "-",
    "The editor still holds the member's code.")
add(F,"Negative","P0","Code with a compile error","Paid member","/careerpilot/practice/:id",
    "1. Submit code with a deliberate syntax error",
    "Remove a semicolon in Java",
    "The COMPILE ERROR TEXT is shown to the student. A bare '(no output)' is a defect — the student cannot learn from it.")
add(F,"Negative","P0","Infinite loop is stopped","Paid member","/careerpilot/practice/:id",
    "1. Submit `while(true){}`\n2. Wait",
    "while True: pass",
    "The run is killed at the timeout and the message says the code ran too long — clearly distinguished from a server error.")
add(F,"Negative","P1","Free member is blocked from practice","Free member","/careerpilot/practice",
    "1. Open practice",
    "-",
    "A clear upgrade prompt, not an error page.")
add(F,"Corner","P0","Many members run heavy code at once","6+ testers, all paid","/careerpilot/practice/:id",
    "1. Have six testers submit Java within the same few seconds",
    "Any correct Java solution",
    "All six eventually succeed. Some may queue. NONE should report a crash or a killed process — six concurrent Java runs previously killed each other, which is why a two-pool queue exists.")
add(F,"Corner","P1","Output differing only by whitespace","Paid member","/careerpilot/practice/:id",
    "1. Submit a correct answer with a trailing newline or extra spaces",
    "-",
    "It is accepted under lenient comparison, with a diff hint if it is rejected. Trailing whitespace must not fail a correct answer.")
add(F,"Corner","P2","Very large output","Paid member","/careerpilot/practice/:id",
    "1. Print 100,000 lines",
    "for i in range(100000): print(i)",
    "Output is truncated with a clear notice. The browser does not hang and the server does not fall over.")

F = "Resume Centre"
add(F,"Positive","P0","Build and save a resume","Paid member","/careerpilot/resume",
    "1. Fill every section\n2. Save\n3. Reload",
    "Realistic fresher details",
    "All content persists exactly as entered.")
add(F,"Positive","P1","AI improve and AI score","Resume saved","/careerpilot/resume",
    "1. Tap Improve\n2. Review suggestions\n3. Apply one\n4. Tap Score",
    "-",
    "Suggestions are specific to the content. Applying one edits only that field. A score with reasoning is returned.")
add(F,"Positive","P2","Import an existing resume","Paid member, PDF/DOCX to hand","/careerpilot/resume",
    "1. Import the file\n2. Check the parsed fields",
    "A real one-page resume",
    "Fields are populated sensibly and anything unparsed is left blank rather than filled with rubbish.")
add(F,"Negative","P1","Import an unsupported file","Paid member","/careerpilot/resume",
    "1. Import a .exe or a 50 MB file",
    "-",
    "Refused with a message naming accepted formats and the size limit. No server error.")
add(F,"Corner","P0","Typing a comma in a list field","Paid member","/careerpilot/resume",
    "1. In Skills, type 'Java, Python, SQL' slowly, character by character",
    "Java, Python, SQL",
    "The comma stays as typed and the list forms correctly. KNOWN PAST DEFECT: the comma was deleted as it was typed. Test this deliberately.")
add(F,"Corner","P2","Very long single field","Paid member","/careerpilot/resume",
    "1. Paste 5,000 characters into a summary field\n2. Save",
    "-",
    "Either accepted or capped with a visible limit. Never a silent truncation.")

F = "AI Mock Interview"
add(F,"Positive","P0","Complete a full mock interview","Paid member, mic permission granted","/careerpilot/interview",
    "1. Start an interview\n2. Allow the microphone\n3. Answer each question aloud\n4. Finish",
    "Speak naturally for 30–60s per answer",
    "Questions are SHORT and asked in an Indian-accented voice. Answers are transcribed. Written feedback is produced at the end AND stored — reopen it later to confirm it persists.")
add(F,"Positive","P1","Review a past interview","At least one interview finished","/careerpilot/interview",
    "1. Open the history list\n2. Open a previous session",
    "-",
    "The full transcript and feedback are readable. Stored feedback is the point of the feature.")
add(F,"Positive","P1","Company-primed interview","A company is live","/careerpilot/companies/:slug → Mock Interview",
    "1. Open a live company\n2. Start its mock interview",
    "-",
    "Questions are recognisably about that company's process, not generic ones.")
add(F,"Negative","P0","Free member cannot start an interview","Free member","/careerpilot/interview",
    "1. Tap Start",
    "-",
    "Blocked with an upgrade prompt quoting the correct price. No AI call is made — this is a cost control, not just a UI state.")
add(F,"Negative","P1","Microphone permission denied","Paid member","/careerpilot/interview",
    "1. Deny the mic prompt\n2. Try to start",
    "-",
    "A clear explanation of why the mic is needed and how to enable it. No blank screen.")
add(F,"Corner","P0","Interview quota is enforced","Paid member with 12 interviews already used","/careerpilot/interview",
    "1. Attempt a 13th interview",
    "-",
    "Should be refused — the plan is 12 per member per year. KNOWN GAP: no quota exists in the code today, so this test WILL FAIL. Raise it as a defect; it is an uncapped cost.")
add(F,"Corner","P1","Silence during an answer","Paid member, interview running","/careerpilot/interview",
    "1. Say nothing for 30 seconds",
    "-",
    "The interview moves on gracefully or prompts the member. It does not hang forever or crash.")
add(F,"Corner","P2","Network drops mid-interview","Paid member, interview running","/careerpilot/interview",
    "1. Disable wifi mid-answer\n2. Re-enable after 20s",
    "-",
    "Either the session recovers or it ends with the transcript so far saved. Work already done must not vanish.")

F = "Prepare Interviews — Company Hub"
add(F,"Positive","P0","Browse live companies","At least one company past the readiness bar","/careerpilot/companies",
    "1. Open the company list",
    "-",
    "Only LIVE companies appear. Companies below the bar must be completely absent, not greyed out.")
add(F,"Positive","P0","All six tabs on a live company","A live company","/careerpilot/companies/:slug",
    "1. Open each tab in turn: Overview, Interview Pattern, Salary, Mock Test, Mock Interview, Questions",
    "Use Cognizant once it is live",
    "Every tab renders real content. No tab is blank and no statistic is shown without its sample size.")
add(F,"Positive","P1","Practise a company question","A live company with questions","/careerpilot/companies/:slug → Questions",
    "1. Open a coding question\n2. Run it",
    "-",
    "It opens in the runnable practice view and executes.")
add(F,"Negative","P0","An unready company 404s by direct URL","A company below the bar","/careerpilot/companies/<unready-slug>",
    "1. Paste the URL of an unready company directly",
    "e.g. /careerpilot/companies/kyndryl",
    "404 / not found. The readiness gate must hold on the direct URL, not only in the list.")
add(F,"Negative","P0","Free member is blocked from the hub","Free member","/careerpilot/companies",
    "1. Open the company list",
    "-",
    "Upgrade prompt. NOTE: the entitlement key company_questions is missing from the production config — verify the lock state renders correctly and raise it if not.")
add(F,"Corner","P0","A company crossing the bar goes live by itself","A company at 19 published questions","/admin/passport/companies",
    "1. As admin, publish a 20th question\n2. As a student, refresh the company list",
    "-",
    "The company appears with no publish action taken. Readiness is computed, never stored.")
add(F,"Corner","P1","A company falling back below the bar","A live company","/admin/passport/companies",
    "1. Unpublish questions until fewer than 20 remain\n2. Refresh as a student",
    "-",
    "The company disappears from the list and its URL 404s again. The gate works in both directions.")
add(F,"Corner","P2","Company names carry no list numbering","-","/careerpilot/companies",
    "1. Read every company name and URL",
    "-",
    "Names read 'Virtusa', not '17. Virtusa'; URLs read /virtusa, not /17-virtusa. KNOWN PAST DEFECT from a pasted numbered list — fixed, but re-check after any bulk import.")

F = "Company Mock Test"
add(F,"Positive","P0","Sit and submit a mock test","Live company with 20+ questions","/careerpilot/companies/:slug → Mock Test",
    "1. Start the test\n2. Answer every question\n3. Submit",
    "-",
    "A score, a pass/fail against the pass mark, and a full review with correct answers and explanations. Questions written for practice are labelled as such.")
add(F,"Positive","P0","Answers survive a refresh","Test in progress","/careerpilot/mock-test/:id",
    "1. Answer five questions\n2. Refresh the browser\n3. Check those five",
    "-",
    "All five answers are still selected and the clock continues from the correct remaining time — it does not restart.")
add(F,"Positive","P1","Resume rather than restart","Test in progress","Leave and return",
    "1. Start a test\n2. Navigate away\n3. Reopen the company and tap Mock Test",
    "-",
    "The SAME attempt resumes. A second paper is not created and no attempt is consumed.")
add(F,"Negative","P0","The clock cannot be cheated","Test in progress","/careerpilot/mock-test/:id",
    "1. Change the computer's system clock backwards\n2. Continue the test",
    "-",
    "The remaining time is unaffected — the deadline is held server-side. The test still auto-submits at the true zero.")
add(F,"Negative","P0","Answer key is not in the page source","Test in progress","/careerpilot/mock-test/:id",
    "1. Open dev tools\n2. Inspect the network response and page source for the correct answers",
    "-",
    "correctIndex and explanations are ABSENT while the test is running. They appear only after submission.")
add(F,"Negative","P1","Attempt limit is enforced","Member who has used all attempts","/careerpilot/companies/:slug",
    "1. Try to start another test",
    "-",
    "Refused with a message naming the attempt limit.")
add(F,"Corner","P0","Time runs out","Test in progress with under a minute left","/careerpilot/mock-test/:id",
    "1. Let the clock reach zero without submitting",
    "-",
    "The test SUBMITS automatically — it must not merely stop. The member is told it was auto-submitted and the answers given are all marked.")
add(F,"Corner","P1","Submitting with blanks","Test in progress","/careerpilot/mock-test/:id",
    "1. Leave several blank\n2. Tap Submit",
    "-",
    "A confirmation names how many are blank. On confirming, blanks are marked wrong and shown as 'left blank' in the review.")
add(F,"Corner","P2","Not enough banked questions","A company with very few questions","/careerpilot/companies/:slug",
    "1. Start a mock test",
    "-",
    "Either the paper is topped up with clearly-labelled generated questions, or a readable message explains there are not enough yet. Never a partial or broken paper.")

F = "Tech News"
add(F,"Positive","P1","Read the news feed","Paid member, items published","/careerpilot/news",
    "1. Open the feed\n2. Open an item",
    "-",
    "Items are listed newest first and open with readable content and a working source link.")
add(F,"Negative","P1","Free member is blocked","Free member","/careerpilot/news",
    "1. Open the feed",
    "-",
    "Upgrade prompt rather than an error.")
add(F,"Corner","P2","Empty feed","No published items","/careerpilot/news",
    "1. Open the feed",
    "-",
    "A friendly empty state, not a blank screen or a spinner that never stops.")

F = "Coins, XP & Leaderboard"
add(F,"Positive","P1","Coins are credited for an activity","Paid member","/careerpilot/coins",
    "1. Note the balance\n2. Complete a mission\n3. Return to the coins screen",
    "-",
    "The balance increases by exactly the admin-configured amount and a ledger line explains why.")
add(F,"Positive","P1","Daily cap is respected","Paid member","/careerpilot/coins",
    "1. Repeat a capped activity past its daily limit",
    "-",
    "Credits stop at the cap. The member is told the cap is reached rather than silently earning nothing.")
add(F,"Positive","P2","Leaderboard reflects real activity","Several members with differing activity","/careerpilot/leaderboard",
    "1. Open the leaderboard",
    "-",
    "Ranking is consistent with actual XP, and the member's own position is visible.")
add(F,"Negative","P1","Coins cannot be spent","Paid member with a balance","/careerpilot/coins",
    "1. Look for any way to redeem or spend",
    "-",
    "EXPECTED TODAY: there is none — the spending half is not built. Confirm the UI does not PROMISE redemption it cannot deliver. Raise anything that implies coins are spendable.")
add(F,"Corner","P1","Replayed activity does not double-credit","Paid member","/careerpilot/coins",
    "1. Complete an earning action\n2. Replay the same request from dev tools",
    "-",
    "Credited exactly once. The ledger's idempotency key must hold.")
add(F,"Corner","P2","Achievements screen with no achievements","New member","/careerpilot/achievements",
    "1. Open achievements",
    "-",
    "A sensible empty state. NOTE: this screen is display-only today with no badge rules behind it.")

F = "Public Passport Card"
add(F,"Positive","P1","Open a card while logged out","A member with a share slug","/careerpilot/card/:slug",
    "1. Copy the card URL\n2. Open it in a private window",
    "-",
    "The card renders fully without a login and shows the member's public achievements.")
add(F,"Negative","P1","An unknown slug","-","/careerpilot/card/doesnotexist",
    "1. Open a made-up slug",
    "-",
    "A clean not-found page, never a server error or a partly-rendered card.")
add(F,"Corner","P2","No private data on the public card","A member with an email and mobile on file","/careerpilot/card/:slug",
    "1. Open the card logged out\n2. Read the page source",
    "-",
    "No mobile number, email or any other private field appears anywhere, including in the page source.")

F = "Plan & Entitlements Config"
add(F,"Positive","P0","Change the membership price","Admin logged in","/admin/passport/config",
    "1. Set the price to 1599\n2. Save\n3. As a student, open a locked feature",
    "priceInr: 1599",
    "Every student-facing prompt and the Razorpay order quote ₹1,599 with no deploy.")
add(F,"Positive","P0","Move a feature between free and paid","Admin logged in","/admin/passport/config",
    "1. Switch one paid feature to free\n2. Save\n3. Check as a FREE member",
    "-",
    "The free member can now use it. Switch it back and access is refused again.")
add(F,"Negative","P1","Invalid price is refused","Admin logged in","/admin/passport/config",
    "1. Enter a negative price or text\n2. Save",
    "-1599 / abcd",
    "Save is refused with a clear message. No invalid value reaches the student-facing prompts.")
add(F,"Corner","P0","All 11 features are listed","Admin logged in","/admin/passport/config",
    "1. Count the entitlement rows",
    "-",
    "Eleven rows including Daily Tech News and Company Interview Questions. KNOWN DEFECT: production holds only 9 — those two are missing and cannot be priced. Expect this to fail today.")

F = "Assessment Builder & Paper Design"
add(F,"Positive","P0","Add a question and see it live","Admin logged in","/admin/passport/assessment",
    "1. Add a question to a stage segment\n2. Save\n3. Start a fresh assessment as a student",
    "-",
    "The new question appears in the correct segment for a new attempt.")
add(F,"Positive","P1","Manage scoring categories","Admin logged in","/admin/passport/assessment",
    "1. Add a scoring category\n2. Attach questions to it\n3. Complete an assessment",
    "-",
    "The category appears in the student's breakdown with a score derived only from its own questions.")
add(F,"Positive","P2","Paper design preview","Admin logged in","/admin/careerpilot/paper-design",
    "1. Change the layout\n2. Preview",
    "-",
    "The preview matches the settings and prints sensibly.")
add(F,"Negative","P1","A question with no correct answer","Admin logged in","/admin/passport/assessment",
    "1. Add options but mark none correct\n2. Save",
    "-",
    "Refused with a clear message. An unmarkable question must never reach a student.")
add(F,"Corner","P1","Deleting a question already answered","A member has answered it","/admin/passport/assessment",
    "1. Delete the question\n2. Open that member's result",
    "-",
    "The historic result still renders. Past answers are not orphaned into a crash.")

F = "Pathways & Missions"
add(F,"Positive","P0","Create a pathway","Admin logged in","/admin/passport/pathways",
    "1. Create a pathway\n2. Save\n3. Assign a member to it",
    "-",
    "It is selectable and the member's content follows it.")
add(F,"Positive","P0","Create a mission for today","Admin logged in","/admin/passport/missions",
    "1. Create a mission for today's date\n2. Save\n3. Open the member dashboard",
    "-",
    "It appears in the member's list for today and can be completed.")
add(F,"Negative","P1","Delete a pathway in use","A member is on that pathway","/admin/passport/pathways",
    "1. Attempt to delete it",
    "-",
    "Either blocked with an explanation, or the affected members are handled explicitly. No member is left pointing at nothing.")
add(F,"Corner","P2","A mission dated in the past","Admin logged in","/admin/passport/missions",
    "1. Create a mission dated last week\n2. Check the member dashboard",
    "-",
    "It does not appear as today's work and does not corrupt the streak.")

F = "Members Admin"
add(F,"Positive","P0","Find a member and read their history","Members exist","/admin/passport/students",
    "1. Search by name, mobile and email\n2. Open the member\n3. Open answers and interviews",
    "-",
    "Search works on all three. Assessment answers and interview transcripts are readable.")
add(F,"Positive","P1","Deactivate and reactivate","An active member","/admin/passport/students",
    "1. Deactivate\n2. Try to log in as that member\n3. Reactivate\n4. Log in again",
    "-",
    "Login is refused while deactivated with a readable message, and works again after reactivation.")
add(F,"Positive","P2","Create a member manually","Admin logged in","/admin/passport/students",
    "1. Create a member with a valid mobile\n2. Save",
    "Mobile: 9876500099",
    "The member is created and can log in via OTP.")
add(F,"Negative","P1","Create with a duplicate mobile","That mobile already exists","/admin/passport/students",
    "1. Create a member on an existing number",
    "-",
    "Refused with a clear duplicate message. No second account is created.")
add(F,"Corner","P1","Deleting a member with history","A member with interviews and submissions","/admin/passport/students",
    "1. Delete them\n2. Open the leaderboard and the ledger",
    "-",
    "Nothing crashes and no orphaned row renders as blank or 'undefined'.")

F = "Company Roster & Content"
add(F,"Positive","P0","Bulk-create companies from a pasted list","Admin logged in","/admin/passport/companies → Roster",
    "1. Paste a NUMBERED list of names\n2. Create",
    "1. TCS\n2. Infosys\n3. Wipro",
    "Companies are created as 'TCS', 'Infosys', 'Wipro' with clean slugs. The numbering must be stripped from BOTH the name and the URL.")
add(F,"Positive","P0","AI-draft all missing profiles","Companies exist without profiles","/admin/passport/companies → Roster",
    "1. Tap Draft all missing\n2. Wait for it to finish",
    "-",
    "Overview, eligibility, salary bands and interview rounds are filled. Eligibility and salary remain UNVERIFIED until a human ticks them.")
add(F,"Positive","P0","Take a company live","A company at 83% needing questions","/admin/passport/companies",
    "1. Import or generate 20+ questions and publish them\n2. Tick Verify eligibility\n3. Check the student list",
    "Use Cognizant — it needs 10 more",
    "The row reaches 100%, and the company appears to students with NO publish button pressed.")
add(F,"Positive","P1","Edit an interview pattern","A company exists","/admin/passport/companies → Interview pattern",
    "1. Add a round, set its duration and what it tests\n2. Save\n3. View as a student",
    "-",
    "The round appears in the student's Interview Pattern tab in the right order.")
add(F,"Positive","P1","Import questions in bulk","A company exists","/admin/passport/companies → Import / Generate",
    "1. Paste a batch of questions\n2. Import\n3. Publish them",
    "-",
    "All are created against the right company and the readiness count rises accordingly.")
add(F,"Negative","P0","Readiness cannot be faked","A company with 19 questions","/admin/passport/companies",
    "1. Verify eligibility and write the overview but leave questions at 19\n2. Check the student list",
    "-",
    "It stays below the bar and remains invisible. Every one of the four checks must pass — no partial credit.")
add(F,"Negative","P1","Draft profile for a nonsense company","Admin logged in","/admin/passport/companies",
    "1. Create a company with a made-up name\n2. Draft its profile",
    "Name: Zzqqxx Technologies",
    "Either it refuses, or it produces obviously-empty content rather than confidently inventing salary bands for a company that does not exist. Invented salary data about a named employer is a serious defect.")
add(F,"Negative","P1","Duplicate company names","A company already exists","/admin/passport/companies",
    "1. Paste a list containing an existing name twice",
    "TCS\nTCS\nInfosys",
    "No duplicates are created. Existing companies are skipped, not overwritten or lost.")
add(F,"Corner","P0","Renaming does not break the URL","A live company","/admin/passport/companies",
    "1. Edit the company name\n2. Save\n3. Open the old student URL",
    "-",
    "The slug does NOT change — existing links keep working. This is deliberate; confirm the name updates while the URL holds.")
add(F,"Corner","P1","Bulleted or dashed pasted list","Admin logged in","/admin/passport/companies → Roster",
    "1. Paste a list using dashes and bullets",
    "- Zoho\n* Freshworks\n• Postman",
    "Names are created as 'Zoho', 'Freshworks', 'Postman' with the markers stripped.")
add(F,"Corner","P2","Questions survive a company edit","A company with questions","/admin/passport/companies",
    "1. Edit the profile, type and tips\n2. Recheck the question count",
    "-",
    "The count is unchanged. Editing a profile must never detach its questions.")

F = "Tech News Admin"
add(F,"Positive","P0","Draft an item from a URL","Admin logged in","/admin/passport/news",
    "1. Paste a public news URL\n2. Tap Draft\n3. Edit and publish",
    "Any public tech news article",
    "A sensible title and summary are drafted, remain editable, and appear in the student feed once published.")
add(F,"Positive","P1","Unpublish an item","A published item","/admin/passport/news",
    "1. Unpublish it\n2. Check the student feed",
    "-",
    "It disappears from the student feed immediately.")
add(F,"Negative","P0","An internal URL is refused","Admin logged in","/admin/passport/news",
    "1. Paste a private or localhost address\n2. Tap Draft",
    "http://localhost:5000/api/v1\nhttp://169.254.169.254/",
    "Refused by the SSRF guard. The server must never fetch its own internal addresses on request — this is a security control, not a convenience.")
add(F,"Negative","P1","An unreachable URL","Admin logged in","/admin/passport/news",
    "1. Paste a URL that 404s",
    "https://example.com/nope-404",
    "A readable failure message. The admin screen does not hang or crash.")
add(F,"Corner","P2","Publishing without a source link","Admin logged in","/admin/passport/news",
    "1. Write an item by hand with no URL\n2. Publish",
    "-",
    "It publishes and the student view renders correctly without a broken or empty link.")

F = "Coins Admin"
add(F,"Positive","P0","Change an earning rule","Admin logged in","/admin/passport/coins",
    "1. Change the coins awarded for one event\n2. Save\n3. Trigger it as a member",
    "-",
    "The new amount is credited. No deploy is needed — every coin number is data.")
add(F,"Positive","P1","Change a daily cap","Admin logged in","/admin/passport/coins",
    "1. Lower a cap to 1\n2. As a member, do the activity twice",
    "-",
    "Only the first is credited.")
add(F,"Positive","P2","Read the ledger","Members have earned coins","/admin/passport/coins → Ledger",
    "1. Open the ledger and filter by member",
    "-",
    "Every credit is listed with its reason and time. The ledger is the audit trail.")
add(F,"Negative","P1","Negative or absurd values","Admin logged in","/admin/passport/coins",
    "1. Set an award to -50 and a cap to -1\n2. Save",
    "-50 / -1",
    "Refused. A negative award would silently drain balances.")
add(F,"Corner","P2","No redemption management exists","Admin logged in","/admin/passport/coins",
    "1. Look for a rewards catalogue or redemption approvals",
    "-",
    "EXPECTED TODAY: none. Confirm nothing in the admin UI implies redemption exists. Raise anything that does.")

F = "Staging Board"
add(F,"Positive","P1","Move a member between stages","A member exists","/admin/careerpilot/staging",
    "1. Move a member to a different stage\n2. Save\n3. Open that member's dashboard",
    "-",
    "The member's stage-driven content changes accordingly.")
add(F,"Negative","P2","Move to an invalid stage","Admin logged in","/admin/careerpilot/staging",
    "1. Attempt an out-of-sequence or unknown stage",
    "-",
    "Refused or handled explicitly, with the member's data left consistent.")

# cross-cutting
F = "Cross-cutting — Access & Security"
add(F,"Negative","P0","Student URLs refuse an anonymous visitor","Logged out","Any /careerpilot/* route",
    "1. Log out\n2. Paste each member URL directly",
    "/careerpilot/roadmap, /practice, /resume, /interview, /companies, /coins",
    "Every one redirects to login. None renders member data, even briefly.")
add(F,"Negative","P0","Admin URLs refuse a student","Logged in as a student","Any /admin/* route",
    "1. Paste each admin URL directly",
    "/admin/passport/config, /students, /companies, /coins",
    "Access is refused. A student must never reach an admin screen or its data.")
add(F,"Negative","P0","One member cannot read another's data","Two members exist","Dev tools",
    "1. As member A, call a detail endpoint using member B's id",
    "Another member's interview or mock-test id",
    "Refused. Every read must be scoped to the calling member.")
add(F,"Corner","P1","Session expiry mid-task","Paid member","Any long task",
    "1. Start a mock test or interview\n2. Leave it until the token expires\n3. Continue",
    "-",
    "The member is asked to log in again cleanly and work already saved is preserved. No blank screen and no silent data loss.")
add(F,"Corner","P1","Exam guards during a mock test","Test in progress","/careerpilot/mock-test/:id",
    "1. Try to select question text and copy it\n2. Try to paste into the page\n3. Try right-click",
    "-",
    "Selection and copy of question text are blocked and pasting is prevented. These guards exist to control cheating — verify they hold on both desktop and mobile.")

F = "Cross-cutting — Mobile"
add(F,"Positive","P0","Full student journey on a phone","A real Android phone","All student routes",
    "1. Join, assess, read the roadmap, practise, run a mock test and a mock interview — all on the phone",
    "Chrome on Android, 360px wide",
    "Every screen is usable. Nothing overflows sideways, no control is unreachable, and the page body never scrolls horizontally.")
add(F,"Positive","P1","Mock test on a small screen","Live company, phone","/careerpilot/mock-test/:id",
    "1. Sit a full test on a phone",
    "-",
    "The question palette, the clock and Submit are all reachable without zooming. The clock stays visible while scrolling.")
add(F,"Corner","P1","Rotating the device mid-test","Test in progress on a phone","/careerpilot/mock-test/:id",
    "1. Rotate to landscape and back",
    "-",
    "Answers and remaining time are unaffected and the layout reflows cleanly.")
add(F,"Corner","P2","Mock interview on mobile data","Paid member on 4G","/careerpilot/interview",
    "1. Run a full interview on mobile data, not wifi",
    "-",
    "Audio plays and answers upload. Slow networks degrade gracefully rather than hanging.")

# ─────────────────────────────────────────────────────────────────────────────
wb = Workbook()

# ── Sheet 1: How to use ──
ws = wb.active
ws.title = "How to use"
widths(ws, [4, 30, 96])
ws["B1"] = "CareerPilot — Manual Test Plan"
ws["B1"].font = Font(bold=True, size=18, color=INK, name="Segoe UI")
ws["B2"] = "End-to-end manual QA for the CareerPilot module, student and admin."
ws["B2"].font = Font(size=11, color="5B6779", name="Segoe UI")
ws["B3"] = "Version 1.1 · 12 August 2026 · commit 24acae87 · 2 defects closed since v1.0"
ws["B3"].font = Font(size=10, color="8E9AAC", name="Segoe UI", italic=True)

rows = [
    ("", ""),
    ("WHAT IS IN THIS FILE", ""),
    ("Feature Guide", "Every feature explained — what it is, why it exists, when a student meets it, "
                      "and where it lives. Read this before writing a single result."),
    ("Test Cases", "The full suite. Filter by Module, Feature, Type or Priority. Fill in Result, "
                   "Tested By, Date and Defect Ref as you go."),
    ("Defect Log", "One row per defect found. Reference the Test Case ID."),
    ("Summary", "Live counts. It reads the Test Cases sheet automatically — do not type into it."),
    ("", ""),
    ("HOW TO RUN A PASS", ""),
    ("1. Set up accounts", "You need FOUR: a brand-new unregistered mobile, a free member, a PAID member, "
                           "and an admin. Most defects hide at the boundary between free and paid."),
    ("2. Work top to bottom", "Cases are ordered as a real student meets them. Running them out of order "
                              "creates preconditions you do not have."),
    ("3. Record what you saw", "If a case fails, write what ACTUALLY happened in Actual Result. "
                               "'Failed' on its own cannot be fixed."),
    ("4. Raise the defect", "Log it in Defect Log and put its reference on the test case row."),
    ("", ""),
    ("PRIORITIES", ""),
    ("P0", "Blocks release. A student cannot complete a core journey, money is at risk, or data leaks."),
    ("P1", "Important. Works but is wrong, confusing, or fails on a common path."),
    ("P2", "Worth fixing. Cosmetic, rare, or an edge case with a workaround."),
    ("", ""),
    ("CASE TYPES", ""),
    ("Positive", "The happy path — the thing works as intended for someone doing it correctly."),
    ("Negative", "Wrong input, wrong state, wrong permission. The system must refuse clearly and safely."),
    ("Corner", "Boundaries and collisions — midnight, expiry, concurrency, a rotated phone, a replayed request. "
               "This is where the real defects live."),
    ("", ""),
    ("KNOWN GAPS — EXPECT THESE TO FAIL", ""),
    ("Price — FIXED 12 Aug", "Production now reads ₹1,599 for 12 months. TC-MEM-002 should PASS; treat a failure as a regression."),
    ("No interview quota", "12 per member per year was specified but never implemented. TC-INT-006 will fail. "
                           "It is an uncapped cost, not a cosmetic issue."),
    ("Coins cannot be spent", "Earning works; spending does not exist. TC-CON-004 documents the expected state."),
    ("Entitlements — FIXED 12 Aug", "All 11 features now registered; Tech News and Company Questions added as paid. TC-CFG-004 should PASS."),
    ("0 companies live", "Until a company clears the readiness bar, every Prepare Interviews and Mock Test "
                         "case is blocked. Take Cognizant live first — it needs 10 more questions."),
    ("Payment never tested", "No membership has ever been purchased in production. Treat TC-MEM-001 as the "
                             "highest-risk case in this document."),
]
r = 5
for k, v in rows:
    if k and not v:
        ws.cell(row=r, column=2, value=k).font = Font(bold=True, size=10, color=HEAD, name="Segoe UI")
    elif k:
        ws.cell(row=r, column=2, value=k).font = Font(bold=True, size=10, color=INK, name="Segoe UI")
        c = ws.cell(row=r, column=3, value=v)
        c.font = Font(size=10, color="374151", name="Segoe UI")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30
    r += 1
ws.sheet_view.showGridLines = False

# ── Sheet 2: Feature Guide ──
ws = wb.create_sheet("Feature Guide")
hdr = ["#", "Module", "Feature", "What it is", "Why it exists", "When a student meets it", "Where it lives"]
ws.append(hdr)
style_header(ws, 1, len(hdr))
widths(ws, [5, 10, 26, 46, 58, 34, 40])
for i, (mod, name, what, why, when, where) in enumerate(FEATURES, start=1):
    ws.append([i, mod, name, what, why, when, where])
    rr = ws.max_row
    for c in range(1, len(hdr) + 1):
        cell = ws.cell(row=rr, column=c)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(size=10, name="Segoe UI")
        cell.border = BORDER
        if i % 2 == 0:
            cell.fill = PatternFill("solid", fgColor=BAND)
    ws.cell(row=rr, column=3).font = Font(size=10, bold=True, name="Segoe UI")
    ws.cell(row=rr, column=7).font = Font(size=9, name="Consolas")
    ws.row_dimensions[rr].height = 62
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:G{ws.max_row}"

# ── Sheet 3: Test Cases ──
ws = wb.create_sheet("Test Cases")
hdr = ["Test Case ID", "Module", "Feature", "Type", "Priority", "Test Case Title",
       "Preconditions", "Navigation Path", "Steps", "Test Data", "Expected Result",
       "Actual Result", "Result", "Tested By", "Date", "Defect Ref"]
ws.append(hdr)
style_header(ws, 1, len(hdr))
widths(ws, [13, 9, 26, 9, 8, 40, 30, 30, 46, 26, 60, 34, 10, 13, 11, 12])

ABBR = {
 "Join, OTP & Login":"JOIN","Career Readiness Assessment":"ASM","Membership & Payment":"MEM",
 "Mission Control & Streak":"MIS","90-Day Roadmap":"ROAD","Practice — Coding, SQL, MCQ":"PRAC",
 "Resume Centre":"RES","AI Mock Interview":"INT","Prepare Interviews — Company Hub":"COMP",
 "Company Mock Test":"TEST","Tech News":"NEWS","Coins, XP & Leaderboard":"CON",
 "Public Passport Card":"CARD","Plan & Entitlements Config":"CFG",
 "Assessment Builder & Paper Design":"ABLD","Pathways & Missions":"PATH","Members Admin":"MADM",
 "Company Roster & Content":"CADM","Tech News Admin":"NADM","Coins Admin":"CIADM",
 "Staging Board":"STG","Cross-cutting — Access & Security":"SEC","Cross-cutting — Mobile":"MOB",
}
MODULE_OF = {f[1]: f[0] for f in FEATURES}
MODULE_OF["Cross-cutting — Access & Security"] = "Both"
MODULE_OF["Cross-cutting — Mobile"] = "Both"

counters = {}
for feature, ttype, prio, title, pre, nav, steps, data, exp in T:
    ab = ABBR[feature]
    counters[ab] = counters.get(ab, 0) + 1
    tcid = f"TC-{ab}-{counters[ab]:03d}"
    ws.append([tcid, MODULE_OF[feature], feature, ttype, prio, title, pre, nav, steps, data, exp,
               "", "", "", "", ""])
    rr = ws.max_row
    fill = {"Positive": POS, "Negative": NEG, "Corner": COR}[ttype]
    for c in range(1, len(hdr) + 1):
        cell = ws.cell(row=rr, column=c)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(size=9.5, name="Segoe UI")
        cell.border = BORDER
    ws.cell(row=rr, column=1).font = Font(size=9.5, bold=True, name="Consolas")
    ws.cell(row=rr, column=4).fill = PatternFill("solid", fgColor=fill)
    ws.cell(row=rr, column=4).font = Font(size=9.5, bold=True, name="Segoe UI")
    ws.cell(row=rr, column=5).alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(row=rr, column=5).font = Font(size=9.5, bold=True, name="Segoe UI",
        color={"P0":"B4232B","P1":"9A5B08","P2":"5B6779"}[prio])
    ws.cell(row=rr, column=6).font = Font(size=9.5, bold=True, name="Segoe UI")
    ws.cell(row=rr, column=8).font = Font(size=9, name="Consolas")
    ws.row_dimensions[rr].height = 78

last = ws.max_row
dv = DataValidation(type="list", formula1='"Pass,Fail,Blocked,Not Run"', allow_blank=True)
ws.add_data_validation(dv); dv.add(f"M2:M{last}")
ws.freeze_panes = "B2"
ws.auto_filter.ref = f"A1:P{last}"

# ── Sheet 4: Defect Log ──
ws = wb.create_sheet("Defect Log")
hdr = ["Defect ID", "Test Case ID", "Module", "Summary", "Steps to Reproduce", "Expected",
       "Actual", "Severity", "Status", "Raised By", "Date", "Notes"]
ws.append(hdr)
style_header(ws, 1, len(hdr))
widths(ws, [11, 13, 10, 42, 46, 36, 36, 10, 12, 13, 11, 30])
seed = [
    ["DEF-001", "TC-MEM-002", "Student", "Membership price is ₹499 in production, not ₹1,599",
     "Open any locked feature as a free member and read the price on the prompt.",
     "₹1,599 everywhere", "₹499 — passportconfigs.priceInr was 499", "Critical", "Fixed", "Audit", "12-Aug-2026",
     "FIXED 12-Aug: priceInr now 1599, confirmed on the live public config. RETEST TC-MEM-002."],
    ["DEF-002", "TC-INT-006", "Student", "No 12-per-year cap on mock interviews",
     "As a paid member, run interviews repeatedly. No limit is applied.",
     "13th attempt refused", "Unlimited interviews allowed", "Critical", "Open", "Audit", "12-Aug-2026",
     "Uncapped AI cost at ~₹2.42 per interview."],
    ["DEF-003", "TC-CON-004", "Student", "Coins can be earned but never spent",
     "Earn coins, then look for any redemption path.",
     "A catalogue or redemption flow", "No spend path exists anywhere", "Major", "Open", "Audit", "12-Aug-2026",
     "No spendCoins function and no redeem endpoint."],
    ["DEF-004", "TC-CFG-004", "Admin", "Entitlements config holds 9 of 11 features",
     "Open /admin/passport/config and count the rows.",
     "11 rows including Tech News and Company Questions", "9 rows — the two newest were absent",
     "Major", "Fixed", "Audit", "12-Aug-2026",
     "FIXED 12-Aug: both appended as paid; no existing tier touched. RETEST TC-CFG-004."],
    ["DEF-005", "TC-COMP-001", "Student", "No company is live, so Prepare Interviews is empty",
     "Open /careerpilot/companies as a paid member.",
     "At least one company listed", "Empty list — all 20 below the readiness bar",
     "Major", "Open", "Audit", "12-Aug-2026", "Content, not code. Cognizant needs 10 more questions."],
]
for row in seed:
    ws.append(row)
    rr = ws.max_row
    for c in range(1, len(hdr) + 1):
        cell = ws.cell(row=rr, column=c)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(size=9.5, name="Segoe UI"); cell.border = BORDER
    ws.cell(row=rr, column=1).font = Font(size=9.5, bold=True, name="Consolas")
    ws.cell(row=rr, column=8).font = Font(size=9.5, bold=True, name="Segoe UI", color="B4232B")
    ws.row_dimensions[rr].height = 56
for extra in range(len(seed) + 2, len(seed) + 40):
    for c in range(1, len(hdr) + 1):
        ws.cell(row=extra, column=c).border = BORDER
dv2 = DataValidation(type="list", formula1='"Critical,Major,Minor,Cosmetic"', allow_blank=True)
ws.add_data_validation(dv2); dv2.add(f"H2:H{len(seed)+40}")
dv3 = DataValidation(type="list", formula1='"Open,In Progress,Fixed,Retest,Closed,Won\'t Fix"', allow_blank=True)
ws.add_data_validation(dv3); dv3.add(f"I2:I{len(seed)+40}")
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:L{len(seed)+40}"

# ── Sheet 5: Summary ──
ws = wb.create_sheet("Summary")
widths(ws, [4, 30, 14, 14, 14, 14, 14, 14])
ws["B1"] = "Execution Summary"
ws["B1"].font = Font(bold=True, size=16, color=INK, name="Segoe UI")
ws["B2"] = "Formulas read the Test Cases sheet. Do not type into the counts."
ws["B2"].font = Font(size=10, italic=True, color="8E9AAC", name="Segoe UI")

R = last  # last row of Test Cases
ws["B4"] = "BY RESULT"; ws["B4"].font = Font(bold=True, size=10, color=HEAD, name="Segoe UI")
res_hdr = ["Total", "Pass", "Fail", "Blocked", "Not Run", "% Executed", "% Passed"]
for i, h in enumerate(res_hdr):
    ws.cell(row=5, column=2 + i, value=h)
style_header(ws, 5, 8)
ws["B6"] = f"='Test Cases'!A1:A{R}" and R - 1
ws["C6"] = f'=COUNTIF(\'Test Cases\'!M2:M{R},"Pass")'
ws["D6"] = f'=COUNTIF(\'Test Cases\'!M2:M{R},"Fail")'
ws["E6"] = f'=COUNTIF(\'Test Cases\'!M2:M{R},"Blocked")'
ws["F6"] = f"=B6-C6-D6-E6"
ws["G6"] = "=IF(B6=0,0,(C6+D6+E6)/B6)"
ws["H6"] = "=IF((C6+D6)=0,0,C6/(C6+D6))"
for col in "BCDEF":
    ws[f"{col}6"].font = Font(bold=True, size=13, name="Consolas")
    ws[f"{col}6"].alignment = Alignment(horizontal="center")
for col in "GH":
    ws[f"{col}6"].number_format = "0.0%"
    ws[f"{col}6"].font = Font(bold=True, size=13, name="Consolas")
    ws[f"{col}6"].alignment = Alignment(horizontal="center")
ws["C6"].font = Font(bold=True, size=13, name="Consolas", color="146B3E")
ws["D6"].font = Font(bold=True, size=13, name="Consolas", color="B4232B")

ws["B9"] = "BY TYPE"; ws["B9"].font = Font(bold=True, size=10, color=HEAD, name="Segoe UI")
for i, h in enumerate(["Type", "Total", "Pass", "Fail"]):
    ws.cell(row=10, column=2 + i, value=h)
style_header(ws, 10, 5)
for i, t in enumerate(["Positive", "Negative", "Corner"]):
    rr = 11 + i
    ws.cell(row=rr, column=2, value=t).font = Font(size=10, bold=True, name="Segoe UI")
    ws.cell(row=rr, column=3, value=f'=COUNTIF(\'Test Cases\'!D2:D{R},B{rr})')
    ws.cell(row=rr, column=4, value=f'=COUNTIFS(\'Test Cases\'!D2:D{R},B{rr},\'Test Cases\'!M2:M{R},"Pass")')
    ws.cell(row=rr, column=5, value=f'=COUNTIFS(\'Test Cases\'!D2:D{R},B{rr},\'Test Cases\'!M2:M{R},"Fail")')

ws["B15"] = "BY PRIORITY"; ws["B15"].font = Font(bold=True, size=10, color=HEAD, name="Segoe UI")
for i, h in enumerate(["Priority", "Total", "Pass", "Fail"]):
    ws.cell(row=16, column=2 + i, value=h)
style_header(ws, 16, 5)
for i, p in enumerate(["P0", "P1", "P2"]):
    rr = 17 + i
    ws.cell(row=rr, column=2, value=p).font = Font(size=10, bold=True, name="Segoe UI")
    ws.cell(row=rr, column=3, value=f'=COUNTIF(\'Test Cases\'!E2:E{R},B{rr})')
    ws.cell(row=rr, column=4, value=f'=COUNTIFS(\'Test Cases\'!E2:E{R},B{rr},\'Test Cases\'!M2:M{R},"Pass")')
    ws.cell(row=rr, column=5, value=f'=COUNTIFS(\'Test Cases\'!E2:E{R},B{rr},\'Test Cases\'!M2:M{R},"Fail")')

ws["B21"] = "BY FEATURE"; ws["B21"].font = Font(bold=True, size=10, color=HEAD, name="Segoe UI")
for i, h in enumerate(["Feature", "Total", "Pass", "Fail"]):
    ws.cell(row=22, column=2 + i, value=h)
style_header(ws, 22, 5)
feats = []
for f in T:
    if f[0] not in feats: feats.append(f[0])
for i, fname in enumerate(feats):
    rr = 23 + i
    ws.cell(row=rr, column=2, value=fname).font = Font(size=10, name="Segoe UI")
    ws.cell(row=rr, column=3, value=f'=COUNTIF(\'Test Cases\'!C2:C{R},B{rr})')
    ws.cell(row=rr, column=4, value=f'=COUNTIFS(\'Test Cases\'!C2:C{R},B{rr},\'Test Cases\'!M2:M{R},"Pass")')
    ws.cell(row=rr, column=5, value=f'=COUNTIFS(\'Test Cases\'!C2:C{R},B{rr},\'Test Cases\'!M2:M{R},"Fail")')
for row in ws.iter_rows(min_row=5, max_row=23 + len(feats), min_col=2, max_col=8):
    for cell in row:
        if cell.value is not None:
            cell.border = BORDER
            if cell.font.size is None or cell.font.size < 11:
                cell.alignment = Alignment(horizontal="center" if cell.column > 2 else "left",
                                           vertical="center")
ws.sheet_view.showGridLines = False

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)

pos = sum(1 for x in T if x[1] == "Positive")
neg = sum(1 for x in T if x[1] == "Negative")
cor = sum(1 for x in T if x[1] == "Corner")
p0  = sum(1 for x in T if x[2] == "P0")
print(f"WROTE {OUT}")
print(f"features described : {len(FEATURES)}")
print(f"test cases         : {len(T)}  (positive {pos} / negative {neg} / corner {cor})")
print(f"P0 cases           : {p0}")
print(f"sheets             : {', '.join(wb.sheetnames)}")
